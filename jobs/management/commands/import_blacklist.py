"""
Import blacklist from Excel file.
Usage: python manage.py import_blacklist <path_to_excel>
Example: python manage.py import_blacklist no_show_candidates.xlsx
"""

from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from jobs.models import Blacklist


class Command(BaseCommand):
    help = "Import blacklist entries from Excel file"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel file")

    def handle(self, *args, **options):
        path = options["file"]
        wb = load_workbook(path, read_only=True)
        ws = wb.active

        created = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Columns: S/N, Name, Phone Number, Role, Comment
            _, name, phone, role, comment = row

            if not name:
                continue

            # Skip duplicates by name
            if Blacklist.objects.filter(name__iexact=str(name).strip()).exists():
                skipped += 1
                continue

            Blacklist.objects.create(
                name=str(name).strip(),
                phone=str(phone).strip() if phone else "",
                role=str(role).strip() if role else "",
                reason=str(comment).strip() if comment else "No show",
            )
            created += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Done — {created} imported, {skipped} skipped (duplicates)."
            )
        )